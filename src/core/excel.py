from io import BytesIO
from typing import Type, TypeVar
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel
from enum import Enum
from loguru import logger

ModelT = TypeVar("ModelT", bound=BaseModel)


class ExcelLoader:

    @staticmethod
    async def load(excel_bytes: bytes, model: Type[ModelT]) -> list[ModelT] | None:
        try:
            wb = load_workbook(BytesIO(excel_bytes), read_only=True)
            ws = wb.active

            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            results: list[ModelT] = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(cell not in (None, "") for cell in row):
                    continue

                row_dict = {}
                for field_name, field_info in model.model_fields.items():
                    if field_name in headers:
                        idx = headers.index(field_name)
                        value = row[idx]
                        row_dict[field_name] = value
                    else:
                        row_dict[field_name] = None
                try:
                    results.append(model(**row_dict))
                except Exception:
                    logger.debug(f"Row: {row_dict} is invalid for model: {model}")

        except Exception as e:
            logger.critical(f"Error loading excel file: {e}")
            return None

        return results

    @staticmethod
    async def save(models: list) -> bytes | None:
        try:
            if not models:
                return None

            wb = Workbook()
            ws = wb.active

            headers = list(models[0].model_fields.keys())
            ws.append(headers)

            for model in models:
                row = []
                for field in headers:
                    value = getattr(model, field)
                    if isinstance(value, Enum):
                        row.append(value.value)
                    else:
                        row.append(value)
                ws.append(row)

            buf = BytesIO()
            wb.save(buf)

            return buf.getvalue()
        except Exception as e:
            logger.critical(f"Error save excel file: {e}")
            return None